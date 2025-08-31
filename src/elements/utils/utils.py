import asyncio
import json
import os
import sys
sys.path.append(os.getcwd())
from src.elements.models.models import client, llm, qwen_llm, chatLLM

def extract_json(s):
    # 查找第一个左大括号的位置，确保最外层是JSON对象
    start = None
    for i, c in enumerate(s):
        if c == '{':
            start = i
            break
    if start is None:
        return None  # 没有找到可能的JSON对象

    stack = ['{']  # 初始化栈，用于匹配括号
    in_string = False  # 是否在字符串内部
    escape = False  # 是否处理转义字符
    end = None  # 记录结束位置

    # 从start+1开始遍历字符，匹配括号
    for i in range(start + 1, len(s)):
        char = s[i]

        # if in_string:
        #     if escape:
        #         escape = False  # 转义状态结束
        #     else:
        #         if char == '\\':
        #             escape = True  # 下一个字符被转义
        #         elif char == '"':
        #             in_string = False  # 字符串结束
        # else:
        # if char == '"':
        #     in_string = True  # 进入字符串
        #     escape = False
        if char == '{':
            stack.append(char)
        elif char == '}':
            if not stack:
                break  # 栈空，无法匹配
            stack.pop()
            if not stack:  # 栈空，匹配成功
                end = i
                break
        # 继续循环直到找到结束或遍历完字符串

    if end is None:
        return None  # 没有完整的结构

    json_str = s[start:end + 1]
    # print(json_str)
    try:
        import json
        parsed = json.loads(json_str)
        # 确保解析结果是一个字典
        return parsed if isinstance(parsed, dict) else None
    except json.JSONDecodeError:
        return None  # JSON解析失败
    
from langsmith import traceable

@traceable(project_name="get_dict")
def get_dict(s):
    s = s.strip('` \njson')
    import json
    try:
        res = json.loads(s)
        return res
    except:
        return extract_json(s)
    

import time
import random
@traceable(project_name="get_completion_with_retry")
def get_completion_with_retry(user_prompt, max_retries=10, initial_delay=1.0):
    retry_count = 0
    
    while retry_count < max_retries:
        try:
            # if trans_model <=50:
            #     completion = client.chat.completions.create(
            #         model="qwen-max-latest",
            #         messages=[
            #             {'role': 'system', 'content': 'you are a helpful assistant.'},
            #             {'role': 'user', 'content': user_prompt}],
            #         # extra_body={"enable_thinking": False}
            #     )
            #     res = get_dict(completion.choices[0].message.content)
            # else:
            model_answer = llm.invoke(user_prompt).content
            
            res = get_dict(model_answer)
                
                
            
            if res is not None:  
                return res
                
        except Exception as e:
            print(f"Attempt {retry_count + 1} failed with error: {str(e)}")
            
            
            # 指数退避 + 随机抖动（避免多个客户端同时重试）
            delay = initial_delay * (2 ** retry_count) + random.uniform(0, 1)
            delay = min(delay, 300)
            print(f"Waiting {delay:.2f} seconds before retry...")
            time.sleep(delay)
            
        retry_count += 1
    
    print(f"Failed to get valid JSON after {max_retries} attempts.")
    return None


count = 0
@traceable(project_name="aget_completion_with_retry")
async def aget_completion_with_retry(user_prompt, max_retries=10, initial_delay=1.0):
    retry_count = 0
    global count
    while retry_count < max_retries:
        try:
            # if trans_model <=50:
            #     completion = client.chat.completions.create(
            #         model="qwen-max-latest",
            #         messages=[
            #             {'role': 'system', 'content': 'you are a helpful assistant.'},
            #             {'role': 'user', 'content': user_prompt}],
            #         # extra_body={"enable_thinking": False}
            #     )
            #     res = get_dict(completion.choices[0].message.content)
            # else:
            model_answer = await chatLLM.ainvoke(user_prompt)
            count += 1
            print(count)
            res = get_dict(model_answer.content)
                
                
            
            if res is not None:  
                return res
                
        except Exception as e:
            print(f"Attempt {retry_count + 1} failed with error: {str(e)}")
            
            
            # 指数退避 + 随机抖动（避免多个客户端同时重试）
            delay = initial_delay * (2 ** retry_count) + random.uniform(0, 1)
            delay = min(delay, 300)
            print(f"Waiting {delay:.2f} seconds before retry...")
            await asyncio.sleep(delay)
            
        retry_count += 1
    
    print(f"Failed to get valid JSON after {max_retries} attempts.")
    return None

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from typing import List

from src.elements.pydmodels.pydmodels import GenerateLinguisticData

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def save_vocabulary_to_excel(data_list, filename="vocabulary_data.xlsx"):
    """
    将词汇数据列表保存到Excel文件的一个工作表中
    参数:
        data_list: 包含多个GenerateLinguisticData对象的列表
        filename: 输出的Excel文件名
    """
    # 创建新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "词汇数据"
    
    # 设置表头
    headers = ["词汇种类", "语言", "词性", "难度", "词汇", "句子"]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    bold_font = Font(bold=True)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 遍历数据并写入Excel
    row = 2
    for data_dict in data_list:
        for category_name, category_data in data_dict.items():
            for language_data in category_data.generate_linguistic_data:
                language = language_data.language
                for vocab_type in language_data.specific_language_vocabularies:
                    word_type = vocab_type.vocabulary_type
                    for difficulty in vocab_type.specific_type_vocabularies:
                        diff_level = difficulty.difficulty
                        for item in difficulty.specific_difficulty_vocabularies:
                            # 写入一行数据
                            ws.cell(row=row, column=1, value=category_name)
                            ws.cell(row=row, column=2, value=language)
                            ws.cell(row=row, column=3, value=word_type)
                            ws.cell(row=row, column=4, value=diff_level)
                            ws.cell(row=row, column=5, value=item.vocabulary)
                            ws.cell(row=row, column=6, value=item.sentence)
                            row += 1
    
    # 设置自动列宽
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 保存文件
    wb.save(filename)
    print(f"数据已成功保存到 {filename}")
    
    
    
import os
import requests
from pathlib import Path
from datetime import datetime, timedelta

def get_upload_policy(api_key, model_name):
    """获取文件上传凭证"""
    url = "https://dashscope.aliyuncs.com/api/v1/uploads"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    params = {
        "action": "getPolicy",
        "model": model_name
    }
    
    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200:
        raise Exception(f"Failed to get upload policy: {response.text}")
    
    return response.json()['data']

def upload_file_to_oss(policy_data, file_path):
    """将文件上传到临时存储OSS"""
    file_name = Path(file_path).name
    key = f"{policy_data['upload_dir']}/{file_name}"
    
    with open(file_path, 'rb') as file:
        files = {
            'OSSAccessKeyId': (None, policy_data['oss_access_key_id']),
            'Signature': (None, policy_data['signature']),
            'policy': (None, policy_data['policy']),
            'x-oss-object-acl': (None, policy_data['x_oss_object_acl']),
            'x-oss-forbid-overwrite': (None, policy_data['x_oss_forbid_overwrite']),
            'key': (None, key),
            'success_action_status': (None, '200'),
            'file': (file_name, file)
        }
        
        response = requests.post(policy_data['upload_host'], files=files)
        if response.status_code != 200:
            raise Exception(f"Failed to upload file: {response.text}")
    
    return f"oss://{key}"

def upload_file_and_get_url(api_key, model_name, file_path):
    """上传文件并获取公网URL"""
    # 1. 获取上传凭证
    policy_data = get_upload_policy(api_key, model_name) 
    # 2. 上传文件到OSS
    oss_url = upload_file_to_oss(policy_data, file_path)
    
    return oss_url

# 使用示例
# if __name__ == "__main__":
    
#     # 从环境变量中获取API Key 或者 在代码中设置 api_key = "your_api_key"
#     api_key = os.getenv("DASHSCOPE_API_KEY")
#     if not api_key:
#         raise Exception("请设置DASHSCOPE_API_KEY环境变量")
        
#     # 设置model名称
#     model_name="qwen-vl-plus"

#     # 待上传的文件路径
#     file_path = "4ba4eef6-6a9b-40ad-932e-8648e4ff6ced-1.png"  # 替换为实际文件路径
    
#     try:
#         public_url = upload_file_and_get_url(api_key, model_name, file_path)
#         expire_time = datetime.now() + timedelta(hours=48)
#         print(f"文件上传成功，有效期为48小时，过期时间: {expire_time.strftime('%Y-%m-%d %H:%M:%S')}")
#         print(f"公网URL: {public_url}")

#     except Exception as e:
#         print(f"Error: {str(e)}")
        
if __name__ == "__main__":
    s = """{
  "RequestId": "A3D6F5CB-7548-538B-9504-4A8DBC09784D",
  "Data": {
    "Status": "PROCESS_SUCCESS",
    "JobId": "D05EB3F3-C33E-5364-93E8-06810C03397E",
    "Result": "{\\"resultUrl\\":\\"https://dashscope-result-sh.oss-cn-shanghai.aliyuncs.com/viapi-video/2025-08-05/88588205-7124-45f3-9239-f131af1bc32f/20250805115528112084_style1_0hlt1cqiw4.jpg?Expires=1754452539&OSSAccessKeyId=LTAI5tQZd8AEcZX6KZV4G8qL&Signature=f%2BA55USci6vLYNOrwy9vzYL1guI%3D\\"}"
  }
}
    """
  
    print(extract_json(s))
    
    
    print(None)